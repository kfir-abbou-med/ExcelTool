from openpyxl.styles import PatternFill

months = {
    1: 'May',
    2: 'June',
    3: 'July',
    4: 'August',
    5: 'September',
    6: 'October',
    7: 'November',
    8: 'December',
    9: 'January',
    10: 'February',
    11: 'March',
    12: 'April',
}

monthsNameToInt = {
    'May': 1,
    'Jun': 2,
    'Jul': 3,
    'Aug': 4,
    'Sep': 5,
    'Oct': 6,
    'Nov': 7,
    'Dec': 8,
    'Jan': 9,
    'Feb': 10,
    'Mar': 11,
    'Apr': 12
}

# cost_centers = {
#     511200: 'R&D',
#     511201: 'QA',
#     511202: 'Finance',
#     511203: 'Admin',
#     511204: 'HR',
#     511205: 'BOD',
#     511206: 'General Mkt',
#     510686: 'R&D General',
#     510801: 'System',
#     510802: 'Cyber Security'
# }
alpha = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'

grand_total_text = 'Grand Total'
comments_text = 'Comments'
sum_of_val_text = 'Sum of Val/COArea Crcy'
cost_center_text = 'Cost Center'
output_file_name = 'P&L by CostCenter.xlsx'
results_text = 'results'
actual_text = 'Actual'
budget_text = 'Budget'
diff_Budget = 'Diff Budget vs Actual'
totals_text = 'Totals'


def get_fill(name):
    white = PatternFill(patternType='solid', fgColor='FFFFFF')
    fill = white
    if name == 'cc':
        fill = PatternFill(patternType='solid', fgColor='f1c232')
    elif name == 'title':
        fill = PatternFill(patternType='solid', fgColor='D9E1F2')
    return fill



