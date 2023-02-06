from openpyxl.utils import get_column_letter

import ExcelUtils


def auto_fit_cols(sheet):
    # iterate through excel
    for j in range(1, sheet.max_column + 1):
        widths_list = []
        for i in range(1, sheet.max_row + 1):
            cell_obj = sheet.cell(row=i, column=j)
            val = str(cell_obj.value)
            if val is not None:
                widths_list.append(len(val))
        col_letter = ExcelUtils.num_hash(j)
        sheet.column_dimensions[str(col_letter)].width = max(widths_list)