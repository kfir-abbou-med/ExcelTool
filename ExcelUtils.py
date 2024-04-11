import Constants
from openpyxl.styles import NamedStyle, Font, Alignment, Border, Side
import re
import datetime
import logging


def get_last_row_column(ws):
    row = ws.max_row
    col = ws.max_column
    return row, col


def get_cell_row_col_with_value(sheet, value):
    for row in sheet.iter_rows(min_row=1, min_col=1):
        for cell in row:
            if cell.value == value:
                print(cell)
                return row, cell
    return 0, 0


def set_absolute_text(sheet, comments_col_num, total_row_num):
    logging.info('[ExcelUtils::set_absolute_text]')
    sheet[f'{num_hash(comments_col_num)}4'] = Constants.comments_text
    sheet[f'A{total_row_num}'] = Constants.grand_total_text
    sheet['C1'] = ''
    sheet['C2'] = ''


def calc_total_for_column(sheet, min_row, max_row, min_col, max_col):
    # logging.info('[ExcelUtils::calc_total_for_column]')
    total_period = 0
    for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if cell.value is not None:
                total_period = float(cell.value) + total_period
        # write_results(sheet, row[0].row, max_col + 1, total_period)
    col_letter = num_hash(min_col)
    month_Number = sheet[f'{col_letter}3'].value

    return (month_Number, total_period)


def copy_data_to_new_sheet(sheet, new_sheet):
    logging.info('[ExcelUtils::copy_data_to_new_sheet]')
    mr = sheet.max_row
    mc = sheet.max_column

    col_offset = new_sheet.max_column

    for i in range(1, mr + 1):
        for j in range(1, mc + 1):
            # reading cell value from source excel file
            c = sheet.cell(row=i, column=j)
            if c.has_style:
                new_sheet.cell(row=i, column=j + col_offset +
                               1)._style = c._style

            # writing the read value to destination excel file
            new_sheet.cell(row=i, column=j + col_offset + 1).value = c.value


def remove_borders(sheet):
    logging.info('[ExcelUtils::remove_borders]')
    any_side = Side(border_style=None)
    border = Border(top=any_side, left=any_side,
                    right=any_side, bottom=any_side)
    for row in sheet.iter_rows(min_row=1, min_col=1):
        for cell in row:
            col = num_hash(cell.column)
            sheet[f'{col}{row[0].row}'].border = border
    return sheet


def set_border_under_row(sheet, min_row, max_row, min_col, max_col):
    logging.info('[ExcelUtils::set_border_under_row]')
    any_side = Side(border_style=None)
    bottom = Side(border_style='thin')
    border = Border(top=any_side, left=any_side, right=any_side, bottom=bottom)
    for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            col = num_hash(cell.column)
            sheet[f'{col}{row[0].row}'].border = border
    return sheet


def set_alignment(sheet, min_row, max_row, min_col, max_col, horizontal, vertical):
    logging.info('[ExcelUtils::set_alignment]')
    for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            col = num_hash(cell.column)
            sheet[f'{col}{row[0].row}'].alignment = Alignment(
                horizontal=horizontal, vertical=vertical)
    return sheet


def set_months_title(sheet, last_col):
    logging.info(f'[ExcelUtils::set_months_title]')
    last_month_int = sheet.cell(row=3, column=last_col).value
    pre_last_month_int = sheet.cell(row=3, column=last_col - 1).value
    last_month_name = Constants.months[last_month_int]
    result_cell_letter = num_hash(last_col+1)

    if str(pre_last_month_int).isnumeric():
        pre_last_month_name = Constants.months[pre_last_month_int]
        sheet[f'{result_cell_letter}4'] = f'{last_month_name} vs {pre_last_month_name}'
    else:
        sheet[f'{result_cell_letter}4'] = f'{last_month_name}'
    return sheet


def is_float(string):
    # logging.info(f'[ExcelUtils::is_float]')
    # Compile a regular expression pattern to match valid float values
    pattern = r"^[-+]?[0-9]*\.?[0-9]+$"
    match = re.match(pattern, string)
    return bool(match)


def calc_months_difference(sheet, min_row, max_row, min_col, max_col):
    logging.info(f'[ExcelUtils::calc_months_difference]')
    for r in range(min_row, max_row):
        # for c in range(max_col-1, max_col):
        current_month_val = sheet.cell(row=r, column=max_col).value
        previous_month_val = sheet.cell(row=r, column=max_col-1).value
        prev_month_text = str(sheet.cell(row=r, column=max_col-1).value)
        is_prev_float = is_float(prev_month_text)

        if prev_month_text is None or not is_prev_float:
            previous_month_val = 0
        if current_month_val is None:
            current_month_val = 0
        sheet.cell(row=r, column=max_col +
                   1).value = current_month_val - previous_month_val
        # set_cell_number_format(sheet.cell(row=r, column=max_col+1))
    return sheet


def set_bold_text(sheet, min_row=1, max_row=None, min_col=1, max_col=None, is_bold=False):
    logging.info(f'[ExcelUtils::set_bold_text]')
    for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            col = num_hash(cell.column)
            sheet[f'{col}{row[0].row}'].font = Font(bold=is_bold)
    return sheet


def set_all_sheet_numbers_to_number_format(sheet, min_row=1, min_col=1):
    logging.info(f'[ExcelUtils::set_all_sheet_numbers_to_number_format]')

    for r in range(min_row, sheet.max_row):
        for c in range(min_col, sheet.max_column):
            cell = sheet.cell(row=r, column=c)
            if cell.value is not None:
                is_numeric = is_float(str(cell.value))
                if is_numeric is True:
                    # set_cell_format_to_currency(cell)
                    set_cell_number_format(cell)


def set_cell_number_format(cell):
    # logging.info(f'[ExcelUtils::set_cell_number_format]')
    num_format = '#,##0.00;"-"#,##0.00'
    cell.number_format = num_format


def set_cell_format_to_currency(cell):
    # logging.info(f'[ExcelUtils::set_cell_number_format]')
    currency_style = NamedStyle(name='currency', number_format='$#,##0.00')
    # Apply the style to the cell
    cell.style = currency_style


def set_fill_on_area(sheet, min_row, max_row, min_col, max_col, color_key):
    logging.info(f'[ExcelUtils::set_fill_on_area]')
    for row in range(min_row, max_row):
        for col in range(min_col, max_col + 1):
            sheet[f'{num_hash(col)}{row}'].fill = Constants.get_fill(color_key)


def set_const_text_sum_sheet(sheet):
    logging.info(f'[ExcelUtils::set_const_text_sum_sheet]')
    # Set titles -> should be called once
    sheet["A1"] = 'Cost Center'
    sheet["A2"] = '(All)'
    sheet["A5"] = Constants.actual_text
    sheet["A6"] = Constants.budget_text
    sheet["A8"] = Constants.diff_Budget  # bold
    last_col = 13
    for i in range(1, last_col):
        col_letter = num_hash(i+1)
        sheet[f'{col_letter}3'] = i
        set_cell_fill(sheet, sheet[f'{col_letter}3'], 'title')
    col_letter = num_hash(last_col + 1)
    sheet[f'{col_letter}3'] = 'Total'
    set_cell_fill(sheet, sheet[f'{col_letter}3'], 'title')


def set_all_totals(sheet, all_sheets_total_per_month):
    logging.info(f'[ExcelUtils::set_all_totals]')
    for key in all_sheets_total_per_month.keys():
        letter = num_hash   (key+1)
        actual = sheet[f'{letter}5']
        budget = sheet[f'{letter}6']
        diff = sheet[f'{letter}8']
        actual.value = all_sheets_total_per_month[key]
        budget.value = 0
        diff.value = f'={letter}6-{letter}5'
        set_cell_number_format(diff)

    sheet["N5"] = '=SUM(B5:M5)'
    sheet["N6"] = '=SUM(B6:M6)'
    sheet["N8"] = '=SUM(B8:M8)'
    set_cell_number_format(sheet["N5"])
    set_cell_number_format(sheet["N6"])
    set_cell_number_format(sheet["N8"])


def set_totals_for_budget(active_sheet, data_sheet, max_row, max_col, all_cost_centers):
    logging.info(f'[ExcelUtils::set_totals_for_budget]')

    min_col = 1
    max_col = 12
    active_sheet_max_row = active_sheet.max_row + 4

    # TODO: set bold
    cost_center_cell = active_sheet[f'A{str(active_sheet_max_row)}']
    cost_center_cell.value = f'{data_sheet.title}- {all_cost_centers[int(data_sheet.title)]}'
    set_cell_bold(active_sheet, active_sheet[f'A{str(active_sheet_max_row)}'])
    set_cell_border(active_sheet, cost_center_cell, False, True, False, False)

    active_sheet[f'A{str(int(active_sheet_max_row+1))}'] = Constants.actual_text
    active_sheet[f'A{str(int(active_sheet_max_row+2))}'] = Constants.budget_text
    active_sheet[f'A{str(int(active_sheet_max_row+3))}'] = Constants.diff_Budget
    set_months_titles(sheet=active_sheet, row=active_sheet_max_row,
                      min_col=2, max_col=14)  # TODO: use args
    min_col = 3

    # set calculated values
    for col in range(min_col, max_col + min_col):
        total_per_month = calc_total_for_column(
            data_sheet, 5, max_row, col, col)
        row_for_results = active_sheet_max_row+1
        col_letter = num_hash(col-1)

        for i in range(min_col, max_col + min_col, 1):
            col_letter = num_hash(i-1)
            actual_cell = active_sheet[f'{col_letter}{row_for_results}']
            budget_cell = active_sheet[f'{col_letter}{str(int(row_for_results+1))}']
            diff_cell = active_sheet[f'{col_letter}{str(int(row_for_results+2))}']

            # check if col is the right period
            month_Short = str(
                active_sheet[f'{col_letter}{row_for_results-1}'].value).split('-')[0]
            month_number = total_per_month[0]
            if month_number != Constants.monthsNameToInt[month_Short]:
                if (actual_cell.value is None):
                    actual_cell.value = 0
                    budget_cell.value = 0
                    diff_cell.value = 0
            else:
                # actual_cell.value = float(total_per_month[1])
                actual_cell.value = total_per_month[1]
                budget_cell.value = 0
                diff_cell.value = f'={col_letter}{str(int(row_for_results+1))}-{col_letter}{str(int(row_for_results))}'
                # set_cell_border(active_sheet, budget_cell,
                #                 False, True, False, False)
                set_cell_number_format(actual_cell)
                set_cell_number_format(budget_cell)
                set_cell_number_format(diff_cell)
                break
            set_cell_border(active_sheet, budget_cell,
                                False, True, False, False)

    # TODO: replace with loop
    actual_total = active_sheet[f'{num_hash(max_col+2)}{str(int(row_for_results))}']
    budget_total = active_sheet[f'{num_hash(max_col + 2)}{str(int(row_for_results + 1))}']
    diff_total = active_sheet[f'{num_hash(max_col+2)}{str(int(row_for_results+2))}']

    actual_total.value = f'=SUM({num_hash(min_col-1)}{row_for_results}:{num_hash(max_col+1)}{row_for_results})'
    budget_total.value = f'=SUM({num_hash(min_col-1)}{row_for_results+1}:{num_hash(max_col+1)}{row_for_results+1})'
    diff_total.value = f'=SUM({num_hash(min_col-1)}{row_for_results+2}:{num_hash(max_col+1)}{row_for_results+2})'

    set_cell_number_format(actual_total)
    set_cell_number_format(budget_total)
    set_cell_number_format(diff_total)


def sum_sheet_total_per_month(sheet, min_row, max_row, min_col, max_col):
    logging.info(f'[ExcelUtils::sum_sheet_total_per_month]')
    total = 0
    for r in sheet.iter_cols(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col):
        for cell in r:
            total = total + float(cell.value)
    return total


def set_cell_fill(sheet, cell, color_key):
    # logging.info(f'[ExcelUtils::set_cell_fill]')
    sheet[cell.coordinate].fill = Constants.get_fill(color_key)


def set_cell_bold(sheet, cell):
    logging.info(f'[ExcelUtils::set_cell_bold]')
    sheet[cell.coordinate].font = Font(bold=True)


def set_cell_border(sheet, cell, top, bottom, left, right):
    # logging.info(f'[ExcelUtils::set_cell_border]')
    no_border = Side(border_style=None)
    thin = Side(border_style='thin')

    top_border = thin if top is True else no_border
    bottom_border = thin if bottom is True else no_border
    left_border = thin if left is True else no_border
    right_border = thin if right is True else no_border

    border = Border(top=top_border, left=left_border,
                    right=right_border, bottom=bottom_border)
    sheet[cell.coordinate].border = border


def set_months_titles(sheet, row, min_col, max_col):
    logging.info(f'[ExcelUtils::set_months_titles]')
    # Set All_cost_center months titles -> should be called once
    for i in range(min_col, max_col):
        letter = num_hash(i)
        month = Constants.months[i-1][:3]
        year = int(str(get_current_year(i))[2:]) - 1
        sheet[f'{letter}{row}'] = f'{month}-{year}'
        set_cell_fill(sheet, sheet[f'{letter}{row}'], 'title')
    letter = num_hash(max_col)
    sheet[f'{letter}{row}'] = 'Total'
    set_cell_fill(sheet, sheet[f'{letter}{row}'], 'title')


def get_current_year(month_key):
    # logging.info(f'[ExcelUtils::get_current_year]')
    if month_key > 9:
        return datetime.date.today().year + 1
    else:
        return datetime.date.today().year


def calc_and_set_total_for_product(sheet, min_row, max_row, min_col, max_col):
    logging.info(f'[ExcelUtils::calc_and_set_total_for_product]')
    total_product = {}
    for row in sheet.iter_rows(min_row, max_row, min_col, max_col):
        for cell in row:
            if cell.value is not None:
                if total_product.__contains__(cell.column):
                    total_product[cell.column] = total_product[cell.column] + cell.value
                else:
                    total_product[cell.column] = cell.value

    for key in total_product.keys():
        col = num_hash(key)
        row = max_row + 1
        sheet[f'{col}{row}'] = total_product[key]
        set_cell_number_format(sheet[f'{col}{row}'])


def write_results(sheet, row, col, total):
    logging.info(f'[ExcelUtils::write_results]')
    col_letter = num_hash(col)
    sheet[f'{col_letter}{row}'] = total


def num_hash(num):
    # logging.info(f'[ExcelUtils::num_hash]')
    alpha = Constants.alpha

    if num < 26:
        return alpha[num-1]
    else:
        q, r = num//26, num % 26
        if r == 0:
            if q == 1:
                return alpha[r-1]
            else:
                return num_hash(q-1) + alpha[r-1]
        else:
            return num_hash(q) + alpha[r-1]
