import sys
import os.path
import pandas as pd
import openpyxl
import AutoFitTool
import Constants
import ExcelUtils
import logging

all_sheets_total_per_month = {}


def read_excel(file):
    # can also index sheet by name or fetch all sheets
    df = pd.read_excel(file, sheet_name=0)
    return df


def set_hard_coded_text(sheet, cost_center, all_cost_centers):
    logging.info("[main::set_hard_coded_text]")
    cost_center_name = all_cost_centers[int(cost_center)]
    sheet["A1"] = Constants.cost_center_text
    sheet["B2"] = cost_center_name + ' $'
    sheet["A3"] = Constants.sum_of_val_text
    sheet['B1'] = cost_center


def get_all_total_per_month(sheet):
    totals = {}
    # Number of columns shifted
    # factor = 2
    for colidx in sheet.iter_cols(min_row=sheet.max_row, min_col=3, max_row=sheet.max_row, max_col=sheet.max_column):
        key = colidx[0].col_idx
        val = sheet[colidx[0].coordinate].value
        month = sheet[f'{colidx[0].column_letter}3'].value
        months_range = range(1, 13, 1)
        if month is not None and month in months_range:
            if val is None:
                val = 0
            # totals[key-factor] = val
            totals[month] = val
    return totals


def init_all_sheets_total_per_month():
    logging.info('[main::init_all_sheets_total_per_month]')
    for i in range(1, 13):
        all_sheets_total_per_month[i] = 0


def add_to_all_sheets_total(single_sheet_total):
    # logging.info('[main::add_to_all_sheets_total]')
    for key in single_sheet_total.keys():
        if all_sheets_total_per_month.__contains__(key):
            all_sheets_total_per_month[key] = all_sheets_total_per_month[key] + \
                single_sheet_total[key]
        else:
            all_sheets_total_per_month[key] = single_sheet_total[key]


def set_temp_sheet_style(curr_sheet, last_row, last_col):
    logging.info(
        f'[main::set_temp_sheet_style] last_row: {last_row}, last_col: {last_col}')
    ExcelUtils.set_fill_on_area(
        curr_sheet, min_row=1, max_row=5, min_col=1, max_col=last_col, color_key='title')
    curr_sheet['B2'].fill = Constants.get_fill('cc')
    ExcelUtils.remove_borders(curr_sheet)
    ExcelUtils.set_border_under_row(
        curr_sheet, last_row, last_row, 1, last_col)
    ExcelUtils.set_border_under_row(
        curr_sheet, 4, 4, last_col + 1, last_col + 2)
    ExcelUtils.set_alignment(curr_sheet, 1, last_row + 1,
                             1, last_col + 1, 'left', 'center')
    ExcelUtils.set_bold_text(sheet=curr_sheet, min_row=1, max_row=last_row + 1, min_col=1, max_col=last_col + 1,
                             is_bold=False)
    ExcelUtils.set_bold_text(sheet=curr_sheet, min_row=last_row + 1, max_row=last_row + 1, min_col=1,
                             max_col=last_col, is_bold=True)
    ExcelUtils.set_months_title(sheet=curr_sheet, last_col=last_col)
    ExcelUtils.calc_months_difference(sheet=curr_sheet, min_row=5, max_row=last_row + 2, min_col=3,
                                      max_col=last_col)
    # ExcelUtils.set_all_sheet_numbers_to_number_format(
    #     curr_sheet, min_row=4, min_col=3)


def set_totals_sheet_style(totals_sheet):
    logging.info('[main::set_temp_sheet_style]')
    ExcelUtils.set_all_totals(totals_sheet, all_sheets_total_per_month)
    ExcelUtils.set_alignment(totals_sheet, 1, totals_sheet.max_row,
                             1, totals_sheet.max_column, 'center', 'center')
    ExcelUtils.set_fill_on_area(
        totals_sheet, 1, 4, 1, totals_sheet.max_column, 'title')
    ExcelUtils.set_bold_text(totals_sheet, 8, 8, 1,
                             totals_sheet.max_column, True)
    ExcelUtils.set_all_sheet_numbers_to_number_format(totals_sheet, min_row=5)


def create_totals_sheet_and_init_consts(workbook, totals):
    logging.info('[main::create_totals_sheet_and_init_consts]')
    workbook.create_sheet(totals)
    totals_sheet = workbook[totals]
    ExcelUtils.set_const_text_sum_sheet(totals_sheet)
    return totals_sheet


def get_all_cost_centers(workbook):
    logging.info('[main::get_all_cost_centers]')
    data = {}
    sheet_name = workbook.sheetnames[0]
    sheet = workbook[sheet_name]
    for row in range(2, sheet.max_row + 1):
        key = int(sheet['B' + str(row)].value)
        value = sheet['C' + str(row)].value

        data[key] = value
    sorted_dict = {k: data[k] for k in sorted(data)}
    logging.info(f'[main::get_all_cost_centers] Sorted_dict: {sorted_dict}')
    return sorted_dict

 # Create and configure logger
logging.basicConfig(filename="log.log", filemode='w', level=logging.DEBUG,
                    format='%(asctime)s - %(levelname)s - %(message)s')


def main_function():
    try:

        logging.info('[main_function] Starting app...')
        excel_dir = r'C:\Temp\ExcelPivotInput'
        
        init_all_sheets_total_per_month()

        input_file = sys.argv[1]  # files[0]
        # input_file = r'C:\Temp\einav\db\db.xlsx'
        # input_file = r'C:\Temp\einav\11-10-23\db_.xlsx'
        # input_file = r'C:\Temp\einav\04-10-24\db_.xlsx'
        logging.info(f'[main_function] input_file: {input_file}')
        print(f'Loaded input: {input_file}')
        if not os.path.exists(excel_dir):
            os.makedirs(excel_dir)

        workbook = openpyxl.load_workbook(input_file)
        cost_centers = get_all_cost_centers(workbook)

        df = read_excel(input_file)

        tmp_output_file = str(f'{excel_dir}\\tmp_out.xlsx')
        pivots = {}

    # Create pivot data and write to file
        with pd.ExcelWriter(tmp_output_file, engine='openpyxl') as writer_pivot:
            for cc in cost_centers:
                logging.info(f'[main_function] Cost center handled: {cc}')

                pivots[cc] = df[df['Cost Center'] == cc].pivot_table(index=['Cost Element', 'Cost element name'],
                                                                     columns=[
                    'Period'],
                    values=[
                    'Val/COArea Crcy'],
                    aggfunc=['sum'])
                center = str(cc)
                center_int = int(cc)
                pivots[center_int].to_excel(writer_pivot, sheet_name=center)

        # load excel file
        workbook = openpyxl.load_workbook(
            filename=tmp_output_file, data_only=False)
        logging.info(f'[main_function] loading workbook')

        totals_sheet = create_totals_sheet_and_init_consts(
            workbook, Constants.totals_text)

        # open workbook
        for sheet in workbook.sheetnames:
            if sheet == Constants.totals_text:
                continue

            curr_sheet = workbook[sheet]
            logging.info(f'[main_function] handling sheet: {sheet}')

            set_hard_coded_text(curr_sheet, sheet, cost_centers)

            last_cell_occupied = ExcelUtils.get_last_row_column(curr_sheet)
            # last_row = last_cell_occupied[0]  # row
            # last_col = last_cell_occupied[1]  # col

            last_row, last_col = last_cell_occupied

            ExcelUtils.set_totals_for_budget(
                totals_sheet, workbook[sheet], last_row, last_col, cost_centers)

            # Set text
            ExcelUtils.calc_and_set_total_for_product(
                curr_sheet, 5, last_row, 3, last_col)
            ExcelUtils.set_absolute_text(
                curr_sheet, last_col + 2, last_cell_occupied[0] + 1)

            # Calculate totals
            totals_for_sheet = get_all_total_per_month(workbook[sheet])
            add_to_all_sheets_total(totals_for_sheet)

            # Set some style issues
            set_temp_sheet_style(curr_sheet, last_row, last_col)

        set_totals_sheet_style(totals_sheet)

        # Copy all results to a single sheet
        results_sheet = workbook.create_sheet(Constants.results_text)
        all_sheets_but_results = (s_ for s_ in workbook.sheetnames if s_ != Constants.results_text and
                                  s_ != Constants.totals_text)

        for sheet in all_sheets_but_results:
            logging.info(f'[main_function] move data to new sheet-> {sheet}')

            active_sheet = workbook[sheet]
            ExcelUtils.copy_data_to_new_sheet(
                sheet=active_sheet, new_sheet=results_sheet)
            # delete sheet
            del workbook[sheet]

        results_sheet.delete_cols(1, 2)

        # saving the destination Excel file
        AutoFitTool.auto_fit_cols(results_sheet)
        AutoFitTool.auto_fit_cols(totals_sheet)
        
        workbook.save(str(Constants.output_file_name))
        dir_path = os.path.dirname(os.path.realpath(__file__))
        print(
            f'Output created in: {dir_path}\\{str(Constants.output_file_name)}')
        logging.info(
            f'Output created in: {dir_path}\\{str(Constants.output_file_name)}')
    except Exception as e:
        logging.error(f'[main_function] Error: {e}')
        exc_type, _, exc_tb = sys.exc_info()
        fname = os.path.split(
        exc_tb.tb_frame.f_code.co_filename)[1]
        print(exc_type, fname, exc_tb.tb_lineno)


main_function()


# Create exe by copying this lines
# pyinstaller --noconfirm --onefile --console --icon "C:/Temp/ExcelPivotInput - Copy/App/images.ico" --distpath "C:\temp\excelToolForEinav\output"  "C:/Users/abbouk2/PycharmProjects/ExcelTool/main.py"
